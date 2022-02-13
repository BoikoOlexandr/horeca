import os.path

import django.db.models as models
from django.db import IntegrityError
from openpyxl import load_workbook
from django.contrib.sites import requests
from django.shortcuts import render, redirect
from django.views import View

from easytrade.forms.upload import UploadFileForm
from easytrade.models import Goods, Orders, Measure, Categories, GoodTypes
from easytrade.views.ViewMixIn import ViewMixIn
from shop.settings import MEDIA_ROOT
from django.template.defaultfilters import slugify as django_slugify

alphabet = {'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i',
            'й': 'j', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't',
            'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch', 'ы': 'i', 'э': 'e', 'ю': 'yu',
            'я': 'ya'}
def slugify(s):
    """
    Overriding django slugify that allows to use russian words as well.
    """
    return django_slugify(''.join(alphabet.get(w, w) for w in s.lower()))


class FillDb(View, ViewMixIn):
    template_name = 'filldb'

    def get(self, request):
        self.set_menu()
        self.content['form'] = UploadFileForm()
        return render(request, self.get_path(), self.content)

    def post(self, request):
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = self.uploaded_file(request.FILES['file'])
            if file:
                self.fill_database(file)
        return render(request, self.get_path(), self.content)

    def uploaded_file(self, file):
        file_extension = file.name.split('.')[-1]
        file_path = os.path.join(MEDIA_ROOT, f'data.{file_extension}')

        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)
        try:
            open(file_path, 'rb')
            return file_path
        except FileExistsError:
            return 0

    def fill_database(self, file):
        result = {}
        sheet = load_workbook(file).active
        row = 1
        while True:
            row += 1
            print(row)
            good = Goods()
            try:
                code = sheet.cell(row=row, column=1).value
                if not code:
                    break
                Goods.objects.get(code=code)
            except good.DoesNotExist:
                good.code = sheet.cell(row=row, column=1).value
            else:
                result[sheet.cell(row=row, column=2).value] = 'is already exist'
                continue
            result[good.code] = []
            good.title = sheet.cell(row=row, column=2).value
            result[good.code].append(good.title)
            good.search_queries = sheet.cell(row=row, column=4).value or self.set_empty_field('search_queries', result[good.code])
            good.search_queries_ukr = sheet.cell(row=row, column=5).value or self.set_empty_field('search_queries_ukr', result[good.code])
            good.description = sheet.cell(row=row, column=7).value or self.set_empty_field('description', result[good.code])
            good.goods_type = self.forigenValue(sheet.cell(row=row, column=8).value, GoodTypes)
            good.price = self.validate_float(sheet.cell(row=row, column=9).value, result[good.code])
            good.measure = self.forigenValue(sheet.cell(row=row, column=11).value, Measure)
            good.category = self.forigenValue(sheet.cell(row=row, column=19).value, Categories)
            good.UID = sheet.cell(row=row, column=25).value or self.set_empty_field('UID', result[good])
            good.good_ID = sheet.cell(row=row, column=26).value or self.set_empty_field('good_ID', result[good.code])
            good.is_published = True
            good.photo = 'defoult.png'
            good.slug = good.code
            good.save()

            result[good.code].append('OK')

        self.content['temp'] = result

    def forigenValue(self, value, model):
        if not value:
            return model.objects.get(name='unknown')
        try:
            instance = model.objects.get(name=value)
        except (model.DoesNotExist, IntegrityError):
            instance = model()
            instance.name = value
            instance.slug = slugify(value)
            instance.save()
        return instance

    def validate_float(self, string: str, result:list) -> float:
        if string:
            string = string.replace(',', '.')
        else:
            string = '0'
            result.append('prise hes no value')
        return float(string)

    def set_empty_field(self, field, result:list):
        result.append(f'{field} hes no value')
        return '0'

